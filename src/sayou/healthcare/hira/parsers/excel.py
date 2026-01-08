# Copyright (c) 2025-2026, Sayouzone
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
엑셀 파일 파싱 모듈

엑셀 파일을 파싱하여 ExcelData 객체로 반환합니다.
"""

import csv
import logging
import os
import openpyxl

from io import BytesIO
from pathlib import Path
from typing import Optional

from ..client import HiraClient
from ..models import ExcelData, DownloadResult
from ..utils import (
    _DOWNLOAD_BASE_URL_,
    decode_euc_kr,
    get_filename,
)

logger = logging.getLogger(__name__)


class ExcelParser:
    """엑셀 파일 파싱 클래스"""

    def __init__(self, client: HiraClient, local_path: str = "./data"):
        """
        ExcelParser 초기화
        
        Args:
            client: HiraClient 인스턴스
            local_path: 로컬 저장 경로
        """
        self._client = client
        self._local_path = local_path

    def parse(self, file_path: str) -> ExcelData:
        """
        엑셀 파일을 파싱하여 ExcelData 객체로 반환
        
        Args:
            file_path: 엑셀 파일 경로
            
        Returns:
            ExcelData: 파싱된 엑셀 데이터
        """
        all_medicines: list[Medicine] = []
        page_results: list[PageResult] = []

        # 현재 디렉토리 설정
        current_dir = Path('./data')

        page_num = 1
        for file in current_dir.glob('의약품등제품정보목록*.xlsx'):
            filename = os.path.join(file_path, file.name)
            excel_data = self.parse_excel_file(filename)

            if excel_data.row_count <= 1:
                return PageResult(
                    page_num=page_num,
                    filename=decoded_filename,
                    download_file=None,
                    medicines=[],
                    has_more=False,
                )

            # Medicine 객체로 변환
            medicines = self._convert_to_medicines(excel_data, page_num, len(excel_data.rows))

            page_result = PageResult(
                page_num=page_num,
                filename=filename,
                download_file=None,
                medicines=medicines,
                has_more=False,
            )

            all_medicines.extend(medicines)
            page_results.append(page_result)

            page_num += 1

        result = DownloadResult(
            medicines=all_medicines,
            page_results=page_results,
            total_pages=page_num + 1,
        )
            
        return result

    def parse_excel_file(self, file_path: str) -> ExcelData:
        """
        엑셀 파일을 파싱하여 ExcelData 객체로 반환
        
        Args:
            file_path: 엑셀 파일 경로
            
        Returns:
            ExcelData: 파싱된 엑셀 데이터
        """
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        sheet_name = sheet.title

        rows = self._parse_excel_sheet(sheet)

        workbook.close()

        filename = Path(file_path).name
        return ExcelData(
            filename=filename,
            rows=rows,
            sheet_name=sheet_name,
        )

    def parse_excel_stream(self, file_stream: bytes, filename: str = "") -> ExcelData:
        """
        엑셀 파일 스트림을 파싱하여 ExcelData 객체로 반환
        
        Args:
            file_stream: 엑셀 파일의 바이트 스트림
            filename: 파일명 (선택)
            
        Returns:
            ExcelData: 파싱된 엑셀 데이터
        """
        workbook = openpyxl.load_workbook(BytesIO(file_stream))
        sheet = workbook.active
        sheet_name = sheet.title

        rows = self._parse_excel_sheet(sheet)

        workbook.close()

        return ExcelData(
            filename=filename,
            rows=rows,
            sheet_name=sheet_name,
        )

    def save_to_csv(
        self,
        excel_data: ExcelData,
        file_path: str,
        columns: Optional[list[str]] = None,
        skip_header_keyword: Optional[str] = None,
        encoding: str = "utf-8-sig",
    ) -> None:
        """
        ExcelData를 CSV 파일로 저장
        
        Args:
            excel_data: 저장할 ExcelData 객체
            file_path: CSV 파일 경로
            columns: 저장할 컬럼명 리스트 (None이면 모든 컬럼)
            skip_header_keyword: 이 키워드가 포함된 행은 건너뜀
            encoding: 파일 인코딩 (기본: utf-8-sig)
        """
        with open(file_path, "w", newline="", encoding=encoding) as file:
            writer = csv.writer(file)

            if columns:
                writer.writerow(columns)

            index = 1
            for row in excel_data.rows:
                if skip_header_keyword and self._row_contains_string(row, skip_header_keyword):
                    continue

                if columns and len(columns) == 3:
                    # id, name, address 형식
                    writer.writerow([index, row[5] if len(row) > 5 else "", row[6] if len(row) > 6 else ""])
                else:
                    writer.writerow(list(row))
                index += 1

    def _parse_excel_sheet(self, sheet) -> list[tuple]:
        """
        엑셀 시트를 파싱하여 튜플 리스트로 반환
        
        Args:
            sheet: openpyxl 시트 객체
            
        Returns:
            파싱된 데이터 (튜플 리스트)
        """
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def _row_contains_string(self, row: tuple, search_string: str) -> bool:
        """
        튜플 내에 특정 문자열이 포함된 요소가 있는지 확인
        
        Args:
            row: 확인할 튜플 (문자열로 구성된 튜플)
            search_string: 찾을 문자열
            
        Returns:
            True: 튜플 내에 search_string을 포함하는 요소가 있는 경우
            False: 그렇지 않은 경우
        """
        for item in row:
            if item and isinstance(item, str) and search_string in item:
                return True
        return False